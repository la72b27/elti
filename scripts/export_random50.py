"""Export 50 randomly sampled records from the full LT MasterList to Excel.

Output: D:\TempStudy\VibeCoding\LS\export_random50.xlsx
"""
from __future__ import annotations
import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

LT_FILE  = r"D:\TempStudy\VibeCoding\LS\Lift Talk - MasterList.xlsx"
OUT_FILE = r"D:\TempStudy\VibeCoding\LS\export_random50.xlsx"
SAMPLE_N = 50
SEED     = None   # set an int (e.g. 42) for reproducible output

COLS = {
    2:  "TC",
    0:  "Town Council",
    3:  "Pfx",
    4:  "Block",
    7:  "Postal Code",
    9:  "Lift Names-All",
    10: "Lift Name - Linked",
    12: "LSS",
    13: "Interface",
    18: "LMD Device ID",
    25: "Full Address",
}
COL_INDICES = sorted(COLS.keys())


def _str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def main() -> None:
    print(f"Reading {LT_FILE} ...")
    wb_in = openpyxl.load_workbook(LT_FILE, read_only=True, data_only=True)
    ws_in = wb_in.active

    all_rows: list[tuple] = []
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        if _str(row[2]):   # skip rows without TC code
            all_rows.append(row)
    wb_in.close()

    print(f"  Total records: {len(all_rows)}")
    n = min(SAMPLE_N, len(all_rows))
    if SEED is not None:
        random.seed(SEED)
    sample = random.sample(all_rows, n)
    print(f"  Sampled {n} records")

    # Write output
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Random 50 Sample"

    hdr_font  = Font(bold=True, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="2A007C")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [COLS[i] for i in COL_INDICES]
    for col_num, label in enumerate(headers, start=1):
        cell = ws_out.cell(row=1, column=col_num, value=label)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align

    row_fill = PatternFill("solid", fgColor="F8F9FA")
    for row_num, raw in enumerate(sample, start=2):
        fill = row_fill if row_num % 2 == 0 else None
        for col_num, col_idx in enumerate(COL_INDICES, start=1):
            val = _str(raw[col_idx]) if col_idx < len(raw) else ""
            cell = ws_out.cell(row=row_num, column=col_num, value=val)
            if fill:
                cell.fill = fill

    for col_num, label in enumerate(headers, start=1):
        max_len = len(label)
        for r in ws_out.iter_rows(min_row=2, max_row=n + 1,
                                   min_col=col_num, max_col=col_num):
            for cell in r:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws_out.column_dimensions[get_column_letter(col_num)].width = min(max_len + 2, 40)

    ws_out.freeze_panes = "A2"
    wb_out.save(OUT_FILE)
    print(f"Saved → {OUT_FILE}  ({n} rows)")


if __name__ == "__main__":
    main()
