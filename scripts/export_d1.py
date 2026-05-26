"""Export 100 random D1 records to Excel (mix of RBE rows + LT-only rows).

Usage:
    python -m scripts.export_d1
Output:
    elti_d1_export_100.xlsx  (in project root)
"""
from __future__ import annotations
import json, random, subprocess, sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT  = ROOT / "elti_d1_export_100.xlsx"
DB   = "elti-db"


# ── wrangler helper ───────────────────────────────────────────────────────────

def wrangler_query(sql: str) -> list[dict]:
    """Run a wrangler d1 execute --remote query; return list of row dicts."""
    cmd = f'npx wrangler d1 execute {DB} --remote --json --command "{sql}"'
    proc = subprocess.run(cmd, capture_output=True, text=True, cwd=ROOT,
                          shell=True)
    if proc.returncode != 0:
        print("[wrangler stderr]", proc.stderr[-400:])
        sys.exit(1)
    # wrangler --json prints non-JSON lines before the JSON array; strip them
    raw = proc.stdout
    start = raw.find("[")
    if start < 0:
        print("ERROR: no JSON in output:", raw[:300])
        sys.exit(1)
    data = json.loads(raw[start:])
    return data[0].get("results", [])


# ── fetch all three tables ────────────────────────────────────────────────────

print("Fetching records table …")
tms_rows = wrangler_query(
    "SELECT tc, pfx, block, lift, address, postcode, lcoy, "
    "status_date, rbe, rbe_display, status FROM records"
)
print(f"  → {len(tms_rows)} TMS rows")

print("Fetching masterlist_lt …")
lt_rows = wrangler_query(
    "SELECT tc, pfx, block, town_council, full_add, postal_code, "
    "lift_names_all, interface, lss, lmd_device_id FROM masterlist_lt"
)
print(f"  → {len(lt_rows)} LT rows")

print("Fetching masterlist_lss …")
lss_rows = wrangler_query(
    "SELECT postal_code, lmd_ip, proxy_ip, vp_tun_ip, lmd_tun_ip, dvr_ip "
    "FROM masterlist_lss"
)
print(f"  → {len(lss_rows)} LSS rows")


# ── build lookup indexes ──────────────────────────────────────────────────────

lt_by_key:  dict[tuple, dict] = {}
lt_by_pc:   dict[str, dict]   = {}
for r in lt_rows:
    key = (r.get("tc",""), r.get("pfx",""), r.get("block",""))
    lt_by_key[key] = r
    pc = (r.get("postal_code") or "").strip()
    if pc:
        lt_by_pc[pc] = r

lss_by_pc: dict[str, dict] = {}
for r in lss_rows:
    pc = (r.get("postal_code") or "").strip()
    if pc:
        lss_by_pc[pc] = r


# ── merge TMS rows with LT + LSS ─────────────────────────────────────────────

def _s(v) -> str:
    return str(v).strip() if v is not None else ""


def merge_tms(row: dict) -> dict:
    key = (_s(row.get("tc")), _s(row.get("pfx")), _s(row.get("block")))
    tms_pc = _s(row.get("postcode"))
    lt  = lt_by_key.get(key) or lt_by_pc.get(tms_pc) or {}
    lt_pc = _s(lt.get("postal_code"))
    eff_pc = lt_pc or tms_pc
    lss = lss_by_pc.get(eff_pc, {})
    return {
        "RBE":          _s(row.get("rbe")),
        "RBE_Display":  _s(row.get("rbe_display")) or _s(row.get("rbe")),
        "TC":           _s(row.get("tc")),
        "Pfx":          _s(row.get("pfx")),
        "Block":        _s(row.get("block")),
        "Lift":         _s(row.get("lift")),
        "Address":      _s(row.get("address")),
        "Postcode":     eff_pc or tms_pc,
        "LCOY":         _s(row.get("lcoy")),
        "Status Date":  _s(row.get("status_date")),
        "Status":       _s(row.get("status")),
        "Town Council": _s(lt.get("town_council")),
        "Full Address": _s(lt.get("full_add")),
        "Lift Names":   _s(lt.get("lift_names_all")),
        "Interface":    _s(lt.get("interface")),
        "LSS":          _s(lt.get("lss")),
        "LMD Device ID":_s(lt.get("lmd_device_id")),
        "LMD IP":       _s(lss.get("lmd_ip")),
        "Proxy IP":     _s(lss.get("proxy_ip")),
        "VP Tun IP":    _s(lss.get("vp_tun_ip")),
        "LMD Tun IP":   _s(lss.get("lmd_tun_ip")),
        "DVR IP":       _s(lss.get("dvr_ip")),
    }


# ── build LT-only rows (no TMS record for this block) ─────────────────────────

tms_keys = set(
    (_s(r.get("tc")), _s(r.get("pfx")), _s(r.get("block")))
    for r in tms_rows
)

lt_only: list[dict] = []
for lt in lt_rows:
    key = (_s(lt.get("tc")), _s(lt.get("pfx")), _s(lt.get("block")))
    if key in tms_keys:
        continue
    pc = _s(lt.get("postal_code"))
    lss = lss_by_pc.get(pc, {})
    lt_only.append({
        "RBE":          "",
        "RBE_Display":  "",
        "TC":           _s(lt.get("tc")),
        "Pfx":          _s(lt.get("pfx")),
        "Block":        _s(lt.get("block")),
        "Lift":         "",
        "Address":      _s(lt.get("full_add")),
        "Postcode":     pc,
        "LCOY":         "",
        "Status Date":  "",
        "Status":       "",
        "Town Council": _s(lt.get("town_council")),
        "Full Address": _s(lt.get("full_add")),
        "Lift Names":   _s(lt.get("lift_names_all")),
        "Interface":    _s(lt.get("interface")),
        "LSS":          _s(lt.get("lss")),
        "LMD Device ID":_s(lt.get("lmd_device_id")),
        "LMD IP":       _s(lss.get("lmd_ip")),
        "Proxy IP":     _s(lss.get("proxy_ip")),
        "VP Tun IP":    _s(lss.get("vp_tun_ip")),
        "LMD Tun IP":   _s(lss.get("lmd_tun_ip")),
        "DVR IP":       _s(lss.get("dvr_ip")),
    })

print(f"  → {len(lt_only)} LT-only (no TMS) rows")

# ── sample 100: 50 TMS + 50 LT-only (or fill from the larger pool) ──────────

merged_tms = [merge_tms(r) for r in tms_rows]

n_tms = min(50, len(merged_tms))
n_lt  = min(50, len(lt_only))
# If one side is short, fill remainder from the other
if n_tms + n_lt < 100:
    if n_tms < 50:
        n_lt = min(100 - n_tms, len(lt_only))
    else:
        n_tms = min(100 - n_lt, len(merged_tms))

sample_tms = random.sample(merged_tms, n_tms)
sample_lt  = random.sample(lt_only, n_lt)
all_rows   = sample_tms + sample_lt
random.shuffle(all_rows)
print(f"Sample: {n_tms} TMS rows + {n_lt} LT-only rows = {len(all_rows)} total")


# ── write Excel ───────────────────────────────────────────────────────────────

HEADERS = [
    "#", "RBE", "RBE Display", "TC", "Pfx", "Block", "Lift", "Address",
    "Postcode", "LCOY", "Status Date", "Status",
    "Town Council", "Full Address", "Lift Names", "Interface", "LSS",
    "LMD Device ID", "LMD IP", "Proxy IP", "VP Tun IP", "LMD Tun IP", "DVR IP",
]
KEYS = [
    None, "RBE", "RBE_Display", "TC", "Pfx", "Block", "Lift", "Address",
    "Postcode", "LCOY", "Status Date", "Status",
    "Town Council", "Full Address", "Lift Names", "Interface", "LSS",
    "LMD Device ID", "LMD IP", "Proxy IP", "VP Tun IP", "LMD Tun IP", "DVR IP",
]

HDR_FILL    = PatternFill("solid", fgColor="2A007C")
HDR_FONT    = Font(color="FFFFFF", bold=True, name="Arial", size=9)
COMF_FILL   = PatternFill("solid", fgColor="EDE0FF")
IOF_FILL    = PatternFill("solid", fgColor="DDF6FF")
LT_FILL     = PatternFill("solid", fgColor="E8F5E9")  # light green for LT-only
ALT_FILL    = PatternFill("solid", fgColor="F8F8F8")
NORMAL_FILL = PatternFill("solid", fgColor="FFFFFF")
CELL_FONT   = Font(name="Arial", size=9)
CENTER      = Alignment(horizontal="center", vertical="center")
LEFT        = Alignment(horizontal="left", vertical="center", wrap_text=False)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "D1 Export"
ws.freeze_panes = "A2"

# Header row
for ci, hdr in enumerate(HEADERS, 1):
    c = ws.cell(1, ci, hdr)
    c.fill = HDR_FILL
    c.font = HDR_FONT
    c.alignment = CENTER

ws.row_dimensions[1].height = 20

# Data rows
for ri, row in enumerate(all_rows, 2):
    rbe = row.get("RBE", "")
    if rbe == "COMF":
        fill = COMF_FILL
    elif rbe == "IOF":
        fill = IOF_FILL
    elif not rbe:
        fill = LT_FILL      # LT-only rows in light green
    else:
        fill = ALT_FILL if ri % 2 == 0 else NORMAL_FILL

    for ci, key in enumerate(KEYS, 1):
        val = (ri - 1) if key is None else (row.get(key) or "")
        c = ws.cell(ri, ci, val)
        c.fill = fill
        c.font = CELL_FONT
        c.alignment = CENTER if ci in (1, 2, 3, 4, 5, 10, 11, 12) else LEFT

# Column widths
COL_WIDTHS = [4, 7, 10, 5, 5, 8, 6, 28, 9, 7, 18, 7,
              18, 32, 28, 12, 8, 16, 14, 14, 14, 14, 14]
for ci, w in enumerate(COL_WIDTHS, 1):
    ws.column_dimensions[get_column_letter(ci)].width = w

# Auto-filter on header
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

# Legend note in row below data
note_row = len(all_rows) + 3
for ci, (label, color) in enumerate([
    ("■ COMF", "9957FF"), ("■ IOF", "22D5FE"), ("■ LT-only (no TMS)", "4CAF50")
], 1):
    c = ws.cell(note_row, ci, label)
    c.font = Font(name="Arial", size=8, color=color, bold=True)

wb.save(OUT)
print(f"\nSaved → {OUT}")
print(f"  Rows: {len(all_rows)}  Columns: {len(HEADERS)}")
print(f"  COMF: {sum(1 for r in all_rows if r['RBE']=='COMF')}  "
      f"IOF: {sum(1 for r in all_rows if r['RBE']=='IOF')}  "
      f"LT-only: {sum(1 for r in all_rows if not r['RBE'])}")
