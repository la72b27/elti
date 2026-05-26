"""Export all LT MasterList records where the same postal_code has multiple
different 'Lift Name - Linked' values (i.e. blocks with more than one LMD device),
enriched with IP data joined from the LSS MasterList.

Output: D:\TempStudy\VibeCoding\LS\export_multi_device.xlsx
"""
from __future__ import annotations
import pandas as pd
from ipaddress import IPv4Address
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from collections import defaultdict

LT_FILE  = r"D:\TempStudy\VibeCoding\LS\Lift Talk - MasterList.xlsx"
LSS_FILE = r"D:\TempStudy\VibeCoding\LS\LSS - MasterList.xlsx"
OUT_FILE = r"D:\TempStudy\VibeCoding\LS\export_multi_device.xlsx"

# LT columns (0-based index → label)
LT_COLS = {
    2:  "TC",
    0:  "Town Council",
    3:  "Pfx",
    4:  "Block",
    7:  "Postal Code",
    9:  "Lift Names-All",
    10: "Lift Name - Linked",
    12: "LSS",
    13: "Interface",
    18: "LMD Device ID",
    25: "Full Address",
}
LT_COL_INDICES = sorted(LT_COLS.keys())

# LSS columns appended after LT columns
LSS_COLS = [
    ("lmd_ip",     "LMD IP"),
    ("proxy_ip",   "Proxy IP"),
    ("vp_tun_ip",  "VP Tun IP"),
    ("lmd_tun_ip", "LMD Tun IP"),
    ("dvr_ip",     "DVR IP"),
]


def _str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _clean_ip(val) -> str:
    s = str(val).strip() if val is not None and str(val).strip() not in ("nan", "") else ""
    return "" if s in ("", "0.0.0.0", "nan", "NaN") else s


def _ip_plus_one(ip_str: str) -> str:
    s = (ip_str or "").strip()
    if not s:
        return ""
    try:
        return str(IPv4Address(int(IPv4Address(s)) + 1))
    except Exception:
        return ""


def _load_lss_ips() -> dict:
    """Return {postal_code: {lmd_ip, proxy_ip, vp_tun_ip, lmd_tun_ip, dvr_ip}}."""
    df = pd.read_excel(LSS_FILE)
    work = df.copy()
    work["postal_code"] = work["postal_code"].fillna(0).astype(float).astype(int).astype(str)
    work = work[work["postal_code"] != "0"]
    work["_is_105"] = (
        work["IP Address"].fillna("").astype(str).str.strip().str.startswith("10.5.")
    )
    result: dict = {}
    for pc, group in work.groupby("postal_code"):
        group = group.sort_values("Lift")
        primary_rows = group[group["_is_105"]]
        primary = primary_rows.iloc[0] if len(primary_rows) else group.iloc[0]
        lmd_ip = _clean_ip(primary.get("IP Address"))
        if len(primary_rows):
            proxy_cands = primary_rows[
                primary_rows["Host2"].apply(lambda v: bool(_clean_ip(v)))
            ]
            proxy_src = proxy_cands.iloc[0] if not proxy_cands.empty else primary
        else:
            proxy_src = primary
        proxy_ip   = _clean_ip(proxy_src.get("Host2"))
        vp_tun_ip  = _clean_ip(proxy_src.get("Gateway1"))
        lmd_tun_ip = _ip_plus_one(vp_tun_ip) if vp_tun_ip else ""
        dvr_parts: list = []
        for _, row in group.iterrows():
            lift = str(row.get("Lift", "")).strip()
            dvr  = _clean_ip(row.get("dvrIP convert"))
            if dvr and lift:
                dvr_parts.append(f"{lift}={dvr}")
        result[str(pc)] = {
            "lmd_ip":      lmd_ip,
            "proxy_ip":    proxy_ip,
            "vp_tun_ip":   vp_tun_ip,
            "lmd_tun_ip":  lmd_tun_ip,
            "dvr_ip":      ", ".join(dvr_parts),
        }
    return result


def main() -> None:
    print(f"Reading {LT_FILE} ...")
    wb_in = openpyxl.load_workbook(LT_FILE, read_only=True, data_only=True)
    ws_in = wb_in.active

    pc_rows: dict[str, list[dict]] = defaultdict(list)
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        tc = _str(row[2])
        if not tc:
            continue
        pc_raw = row[7]
        if pc_raw is None:
            continue
        try:
            pc = str(int(float(pc_raw)))
        except (ValueError, TypeError):
            pc = _str(pc_raw)
        if not pc or pc == "0":
            continue
        linked = _str(row[10])
        pc_rows[pc].append({"row": row, "linked": linked})
    wb_in.close()

    multi: list[tuple[str, list]] = []
    for pc, entries in sorted(pc_rows.items()):
        if len({e["linked"] for e in entries}) > 1:
            multi.append((pc, entries))

    print(f"  Found {len(multi)} postal codes with multiple Lift Name - Linked values "
          f"({sum(len(e) for _, e in multi)} total rows)")

    print(f"Reading {LSS_FILE} ...")
    lss_ips = _load_lss_ips()
    print(f"  Loaded {len(lss_ips)} LSS records")

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Multi-Device Blocks"

    hdr_font  = Font(bold=True, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="2A007C")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    lt_headers  = [LT_COLS[i] for i in LT_COL_INDICES]
    lss_headers = [label for _, label in LSS_COLS]
    headers     = lt_headers + lss_headers
    total_cols  = len(headers)

    for col_num, label in enumerate(headers, start=1):
        cell = ws_out.cell(row=1, column=col_num, value=label)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align

    fill_a = PatternFill("solid", fgColor="F0EEFF")
    fill_b = PatternFill("solid", fgColor="FFFFFF")

    row_num = 2
    for idx, (pc, entries) in enumerate(multi):
        grp_fill = fill_a if idx % 2 == 0 else fill_b
        lss_row  = lss_ips.get(pc, {})
        for entry in entries:
            raw = entry["row"]
            for col_num, col_idx in enumerate(LT_COL_INDICES, start=1):
                val = _str(raw[col_idx]) if col_idx < len(raw) else ""
                cell = ws_out.cell(row=row_num, column=col_num, value=val)
                cell.fill = grp_fill
            for offset, (key, _lbl) in enumerate(LSS_COLS):
                col_num = len(LT_COL_INDICES) + offset + 1
                val = lss_row.get(key, "")
                cell = ws_out.cell(row=row_num, column=col_num, value=val)
                cell.fill = grp_fill
            row_num += 1

    for col_num in range(1, total_cols + 1):
        max_len = len(headers[col_num - 1])
        for r in ws_out.iter_rows(min_row=2, max_row=row_num - 1,
                                   min_col=col_num, max_col=col_num):
            for cell in r:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws_out.column_dimensions[get_column_letter(col_num)].width = min(max_len + 2, 40)

    ws_out.freeze_panes = "A2"
    wb_out.save(OUT_FILE)
    print(f"Saved → {OUT_FILE}  ({row_num - 2} rows, {len(multi)} blocks, {total_cols} columns)")


if __name__ == "__main__":
    main()
