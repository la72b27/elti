"""Upload Lift Talk MasterList to ELTI Worker D1 (masterlist_lt table).

Usage:
    set ELTI_UPDATE_TOKEN=<your-token>
    python -m scripts.upload_lt

The script reads D:\TempStudy\VibeCoding\LS\Lift Talk - MasterList.xlsx,
extracts the required columns, and POSTs to the Worker's /api/lt/upload
endpoint in batches of 200.

Required env vars:
    ELTI_UPDATE_TOKEN   Worker auth token
    ELTI_WORKER_URL     (optional, default: https://elti.insg.vip)
"""

from __future__ import annotations
import os, json, sys
import httpx
import openpyxl

LT_FILE     = r"D:\TempStudy\VibeCoding\LS\Lift Talk - MasterList.xlsx"
WORKER_URL  = os.environ.get("ELTI_WORKER_URL", "https://elti.insg.vip").rstrip("/")
AUTH_TOKEN  = os.environ.get("ELTI_UPDATE_TOKEN", "")
BATCH_SIZE  = 200

# Column mapping in the Excel sheet (0-based indices from headers)
# Town_Council(0), town_council_code(2), Pre(3), block(4),
# postal_code(7), Lift Names-All(9), LSS(12), Interface(13),
# LMD Device ID(18), Full Add(25)
_WANT = {
    "Town_Council":      0,
    "town_council_code": 2,
    "Pre":               3,
    "block":             4,
    "postal_code":       7,
    "Lift Names-All":    9,
    "LSS":               12,
    "Interface":         13,
    "LMD Device ID":     18,
    "Full Add":          25,
}


def _str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def read_lt() -> list[dict]:
    wb = openpyxl.load_workbook(LT_FILE, read_only=True, data_only=True)
    ws = wb.active

    # Verify headers match expected indices
    raw_headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    for name, idx in _WANT.items():
        actual = raw_headers[idx] if idx < len(raw_headers) else None
        if actual != name:
            print(f"  [warn] col[{idx}] expected '{name}', got '{actual}' — proceeding anyway")

    records: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        tc = _str(row[_WANT["town_council_code"]])
        if not tc:
            continue
        records.append({
            "tc":            tc,
            "pfx":           _str(row[_WANT["Pre"]]),
            "block":         _str(row[_WANT["block"]]),
            "town_council":  _str(row[_WANT["Town_Council"]]),
            "full_add":      _str(row[_WANT["Full Add"]]),
            "postal_code":   _str(row[_WANT["postal_code"]]),
            "lift_names_all":_str(row[_WANT["Lift Names-All"]]),
            "interface":     _str(row[_WANT["Interface"]]),
            "lss":           _str(row[_WANT["LSS"]]),
            "lmd_device_id": _str(row[_WANT["LMD Device ID"]]),
        })

    wb.close()
    return records


def upload(records: list[dict]) -> None:
    if not AUTH_TOKEN:
        print("ERROR: ELTI_UPDATE_TOKEN not set"); sys.exit(1)

    headers = {
        "Content-Type":  "application/json",
        "X-Update-Token": AUTH_TOKEN,
    }
    total = len(records)
    upserted = 0
    with httpx.Client(timeout=60) as client:
        for i in range(0, total, BATCH_SIZE):
            batch = records[i:i + BATCH_SIZE]
            resp  = client.post(
                f"{WORKER_URL}/api/lt/upload",
                content=json.dumps({"records": batch}),
                headers=headers,
            )
            resp.raise_for_status()
            n = resp.json().get("upserted", len(batch))
            upserted += n
            print(f"  batch {i // BATCH_SIZE + 1}/{-(-total // BATCH_SIZE)}: "
                  f"{n} records upserted (cumulative {upserted}/{total})")


def main() -> None:
    print(f"Reading {LT_FILE} ...")
    records = read_lt()
    print(f"  Loaded {len(records)} records from Lift Talk MasterList")
    print(f"Uploading to {WORKER_URL} ...")
    upload(records)
    print("Done.")


if __name__ == "__main__":
    main()
