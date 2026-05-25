"""Generate lt_data.sql from Lift Talk MasterList Excel."""
import openpyxl

LT_FILE  = r"D:\TempStudy\VibeCoding\LS\Lift Talk - MasterList.xlsx"
OUT_FILE = r"D:\TempStudy\VibeCoding\ELTI\migrations\lt_data.sql"

COL = {
    "Town_Council": 0, "town_council_code": 2, "Pre": 3, "block": 4,
    "postal_code": 7, "Lift_Names_All": 9, "LSS": 12,
    "Interface": 13, "Full_Add": 25
}

def esc(v):
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip().replace("'", "''")

wb = openpyxl.load_workbook(LT_FILE, read_only=True, data_only=True)
ws = wb.active
next(ws.iter_rows(min_row=1, max_row=1))  # skip header

lines = ["DELETE FROM masterlist_lt;"]
count = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    tc = esc(row[COL["town_council_code"]])
    if not tc:
        continue
    pfx          = esc(row[COL["Pre"]])
    block        = esc(row[COL["block"]])
    town_council = esc(row[COL["Town_Council"]])
    full_add     = esc(row[COL["Full_Add"]])
    postal_code  = esc(row[COL["postal_code"]])
    lift_names   = esc(row[COL["Lift_Names_All"]])
    interface    = esc(row[COL["Interface"]])
    lss          = esc(row[COL["LSS"]])
    lines.append(
        "INSERT OR REPLACE INTO masterlist_lt "
        "(tc,pfx,block,town_council,full_add,postal_code,lift_names_all,interface,lss) "
        "VALUES "
        f"('{tc}','{pfx}','{block}','{town_council}','{full_add}',"
        f"'{postal_code}','{lift_names}','{interface}','{lss}');"
    )
    count += 1
wb.close()

sql = "\n".join(lines)
with open(OUT_FILE, "w", encoding="utf-8") as f:
    f.write(sql)
print(f"Written {count} rows → {OUT_FILE}  ({len(sql)//1024} KB)")
